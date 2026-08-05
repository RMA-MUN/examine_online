from datetime import datetime, timedelta
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import AsyncMock

import pytest
from openpyxl import load_workbook

from app.services.dashboard_export_service import (
    DashboardExportError,
    allowed_datasets_for_role,
    get_dashboard_export_datasets,
    render_dashboard_export,
)
from app.api import statistics
from fastapi import HTTPException
from fastapi.testclient import TestClient
from fastapi.responses import StreamingResponse


class ScalarResult:
    def __init__(self, values):
        self.values = values

    def scalars(self):
        return self

    def all(self):
        return self.values

    def scalar_one_or_none(self):
        return self.values

    def scalar_one(self):
        return self.values


class RowResult:
    def __init__(self, rows):
        self.rows = rows

    def all(self):
        return self.rows


class FakeSession:
    def __init__(self, responses):
        self.execute = AsyncMock(side_effect=responses)


@pytest.fixture
def student_user():
    return SimpleNamespace(id=1, role="student")


@pytest.fixture
def teacher_user():
    return SimpleNamespace(id=2, role="teacher")


@pytest.fixture
def admin_user():
    return SimpleNamespace(id=3, role="admin")


@pytest.fixture
def student_db():
    now = datetime.now()
    student_record = SimpleNamespace(
        id=1,
        student_id=1,
        exam_id=20,
        score=80,
        status="graded",
        start_time=now - timedelta(days=2),
        submit_time=now - timedelta(days=1),
    )
    other_student_record = SimpleNamespace(
        id=2,
        student_id=99,
        exam_id=21,
        score=100,
        status="graded",
        start_time=now,
        submit_time=now,
    )
    available_exam = SimpleNamespace(
        id=30,
        title="Upcoming exam",
        start_time=now + timedelta(days=1),
        end_time=now + timedelta(days=2),
        duration=90,
        status="published",
    )
    return FakeSession(
        [
            ScalarResult([available_exam]),
            ScalarResult([student_record, other_student_record]),
            RowResult([(student_record, 60), (other_student_record, 60)]),
            ScalarResult([student_record, other_student_record]),
            ScalarResult(60),
            ScalarResult("Owned exam"),
        ]
    )


@pytest.fixture
def teacher_db():
    now = datetime.now()
    owned_courses = [
        SimpleNamespace(id=10, teacher_id=2),
        SimpleNamespace(id=11, teacher_id=2),
        SimpleNamespace(id=12, teacher_id=999),
    ]
    owned_exam = SimpleNamespace(
        id=20,
        course_id=10,
        title="Owned exam",
        status="published",
        start_time=now,
    )
    foreign_exam = SimpleNamespace(
        id=21,
        course_id=12,
        title="Foreign exam",
        status="published",
        start_time=now + timedelta(days=1),
    )
    owned_record = SimpleNamespace(id=30, exam_id=20)
    foreign_record = SimpleNamespace(id=31, exam_id=21)
    return FakeSession(
        [
            ScalarResult(owned_courses),
            ScalarResult([owned_exam, foreign_exam]),
            ScalarResult([owned_record, foreign_record]),
            ScalarResult(
                [
                    SimpleNamespace(record_id=30, grading_source="pending"),
                    SimpleNamespace(record_id=30, grading_source="pending"),
                    SimpleNamespace(record_id=31, grading_source="pending"),
                ]
            ),
        ]
    )


@pytest.fixture
def admin_db():
    now = datetime.now()
    users = [
        SimpleNamespace(id=1, username="student", name="Student", role="student", created_at=now),
        SimpleNamespace(id=2, username="teacher-a", name="Teacher A", role="teacher", created_at=now),
        SimpleNamespace(id=3, username="teacher-b", name="Teacher B", role="teacher", created_at=now),
        SimpleNamespace(id=4, username="admin", name="Admin", role="admin", created_at=now),
        SimpleNamespace(id=5, username="student-b", name="Student B", role="student", created_at=now),
        SimpleNamespace(id=6, username="student-c", name="Student C", role="student", created_at=now),
        SimpleNamespace(id=7, username="student-d", name="Student D", role="student", created_at=now),
    ]
    return FakeSession([ScalarResult(users), ScalarResult(5)])


@pytest.mark.asyncio
async def test_student_export_contains_only_student_records(student_db, student_user):
    datasets = await get_dashboard_export_datasets(student_db, student_user)

    assert set(datasets) == {"summary", "recent_records", "upcoming_exams"}
    assert all(row["student_id"] == student_user.id for row in datasets["recent_records"])


@pytest.mark.asyncio
async def test_teacher_export_is_limited_to_owned_courses(teacher_db, teacher_user):
    datasets = await get_dashboard_export_datasets(teacher_db, teacher_user)

    assert datasets["pending_grading"] == [{"exam_title": "Owned exam", "pending_count": 2}]
    assert all(row["course_id"] in {10, 11} for row in datasets["recent_exams"])


@pytest.mark.asyncio
async def test_admin_export_contains_global_role_distribution(admin_db, admin_user):
    datasets = await get_dashboard_export_datasets(admin_db, admin_user)

    assert datasets["role_distribution"] == [
        {"role": "student", "count": 4},
        {"role": "teacher", "count": 2},
        {"role": "admin", "count": 1},
    ]


def test_allowed_datasets_are_scoped_to_known_roles():
    assert allowed_datasets_for_role("student") == frozenset(
        {"summary", "recent_records", "upcoming_exams"}
    )
    assert allowed_datasets_for_role("teacher") == frozenset(
        {"summary", "pending_grading", "recent_exams"}
    )
    assert allowed_datasets_for_role("admin") == frozenset(
        {"summary", "role_distribution", "recent_users"}
    )
    assert allowed_datasets_for_role("unknown") == frozenset()


def test_csv_renderer_writes_utf8_bom_and_headers():
    content, media_type, filename = render_dashboard_export(
        {"summary": [{"metric": "pass_rate", "value": 80}]}, "csv", "summary"
    )

    assert content.startswith(b"\xef\xbb\xbf")
    assert media_type == "text/csv; charset=utf-8"
    assert filename.endswith(".csv")
    assert "指标" in content.decode("utf-8-sig")


def test_csv_renderer_rejects_unknown_dataset():
    with pytest.raises(DashboardExportError, match="Unknown dataset"):
        render_dashboard_export({"summary": []}, "csv", "not_allowed")


def test_xlsx_renderer_creates_expected_sheets():
    content, media_type, filename = render_dashboard_export(
        {"summary": [{"metric": "exam_count", "value": 3}], "role_distribution": []}, "xlsx"
    )

    assert media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert filename.endswith(".xlsx")
    workbook = load_workbook(BytesIO(content), read_only=True)
    assert workbook.sheetnames == ["summary", "role_distribution"]
    assert list(workbook["role_distribution"].values) == [("角色", "数量")]


@pytest.mark.asyncio
async def test_student_can_download_xlsx_with_attachment_headers(
    monkeypatch, student_db, student_user
):
    datasets = {"summary": [{"metric": "pass_rate", "value": 80}]}
    monkeypatch.setattr(
        statistics,
        "get_dashboard_export_datasets",
        AsyncMock(return_value=datasets),
    )

    response = await statistics.export_dashboard_file(
        file_format="xlsx",
        dataset=None,
        db=student_db,
        current_user=student_user,
    )

    assert isinstance(response, StreamingResponse)
    assert response.media_type == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert response.headers["content-disposition"].startswith("attachment;")


@pytest.mark.asyncio
async def test_student_cannot_request_teacher_dataset(student_db, student_user):
    with pytest.raises(HTTPException) as exc_info:
        await statistics.export_dashboard_file(
            file_format="csv",
            dataset="pending_grading",
            db=student_db,
            current_user=student_user,
        )

    assert exc_info.value.status_code == 400


def test_cors_exposes_content_disposition_header():
    from app.main import app

    with TestClient(app) as client:
        response = client.options(
            "/api/statistics/dashboard/export?format=csv",
            headers={
                "Origin": "http://localhost:3000",
                "Access-Control-Request-Method": "GET",
            },
        )

    assert response.status_code == 200
    exposed_headers = response.headers.get("access-control-expose-headers", "")
    assert "content-disposition" in exposed_headers.lower()


@pytest.mark.asyncio
async def test_export_rejects_invalid_format(student_db, student_user):
    with pytest.raises(HTTPException) as exc_info:
        await statistics.export_dashboard_file(
            file_format="pdf",
            dataset=None,
            db=student_db,
            current_user=student_user,
        )

    assert exc_info.value.status_code == 400
