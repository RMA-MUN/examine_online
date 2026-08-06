from io import BytesIO
from types import SimpleNamespace
from unittest.mock import AsyncMock

import pytest
from fastapi import HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook

from app.api import statistics
from app.services.score_export_service import (
    ScoreExportError,
    build_score_export_data,
    get_score_export_options,
    render_score_export,
)


class ScalarResult:
    def __init__(self, values):
        self.values = values

    def scalars(self):
        return self

    def all(self):
        return self.values


class RowResult:
    def __init__(self, rows):
        self.rows = rows

    def all(self):
        return self.rows


class FakeSession:
    def __init__(self, responses):
        self.execute = AsyncMock(side_effect=responses)


def make_record(rid, sid, eid, score, status):
    return SimpleNamespace(id=rid, student_id=sid, exam_id=eid, score=score, status=status)


@pytest.mark.asyncio
async def test_admin_build_aggregates_class_course_exam_summary():
    class_a = SimpleNamespace(id=1, name="计科2401班")
    class_b = SimpleNamespace(id=2, name="计科2402班")
    student_a = SimpleNamespace(id=10, name="小明", class_id=1)
    student_b = SimpleNamespace(id=11, name="小红", class_id=1)
    student_c = SimpleNamespace(id=12, name="小刚", class_id=2)
    course = SimpleNamespace(id=1, name="数学")
    exam = SimpleNamespace(id=1, course_id=1, title="期中考试", pass_score=60)
    rows = [
        (make_record(1, 10, 1, 80, "graded"), student_a, course, exam, class_a),
        (make_record(2, 11, 1, 55, "submitted"), student_b, course, exam, class_a),
        (make_record(3, 12, 1, 90, "graded"), student_c, course, exam, class_b),
    ]
    db = FakeSession([RowResult(rows), ScalarResult([])])

    data = await build_score_export_data(db, SimpleNamespace(id=3, role="admin"))

    assert data["class_summary"] == [
        {
            "class_name": "计科2401班",
            "course_name": "数学",
            "exam_title": "期中考试",
            "student_count": 2,
            "avg_score": 67.5,
            "pass_rate": 50.0,
            "max_score": 80,
            "min_score": 55,
        },
        {
            "class_name": "计科2402班",
            "course_name": "数学",
            "exam_title": "期中考试",
            "student_count": 1,
            "avg_score": 90.0,
            "pass_rate": 100.0,
            "max_score": 90,
            "min_score": 90,
        },
    ]
    assert data["student_scores"] == [
        {
            "student_name": "小明",
            "class_name": "计科2401班",
            "course_name": "数学",
            "exam_title": "期中考试",
            "score": 80,
            "pass_score": 60,
            "status": "已阅卷",
        },
        {
            "student_name": "小红",
            "class_name": "计科2401班",
            "course_name": "数学",
            "exam_title": "期中考试",
            "score": 55,
            "pass_score": 60,
            "status": "待阅卷",
        },
        {
            "student_name": "小刚",
            "class_name": "计科2402班",
            "course_name": "数学",
            "exam_title": "期中考试",
            "score": 90,
            "pass_score": 60,
            "status": "已阅卷",
        },
    ]


@pytest.mark.asyncio
async def test_build_marks_students_without_class():
    course = SimpleNamespace(id=1, name="数学")
    exam = SimpleNamespace(id=1, course_id=1, title="期中考试", pass_score=60)
    student = SimpleNamespace(id=10, name="小明", class_id=None)
    db = FakeSession(
        [
            RowResult(
                [(make_record(1, 10, 1, 80, "graded"), student, course, exam, None)]
            ),
            ScalarResult([]),
        ]
    )

    data = await build_score_export_data(db, SimpleNamespace(id=3, role="admin"))

    assert data["class_summary"][0]["class_name"] == "未分配班级"
    assert data["student_scores"][0]["class_name"] == "未分配班级"


@pytest.mark.asyncio
async def test_teacher_build_filters_to_assigned_courses():
    math = SimpleNamespace(id=1, name="数学")
    exam_math = SimpleNamespace(id=1, course_id=1, title="期中考试", pass_score=60)
    class_a = SimpleNamespace(id=1, name="一班")
    student = SimpleNamespace(id=10, name="小明", class_id=1)
    db = FakeSession(
        [
            ScalarResult([1]),
            RowResult(
                [
                    (make_record(1, 10, 1, 80, "graded"), student, math, exam_math, class_a),
                ]
            ),
            ScalarResult([]),
        ]
    )

    data = await build_score_export_data(db, SimpleNamespace(id=2, role="teacher"))

    assert [row["course_name"] for row in data["student_scores"]] == ["数学"]


@pytest.mark.asyncio
async def test_teacher_build_rejects_course_outside_assignment():
    db = FakeSession([ScalarResult([1, 3])])

    with pytest.raises(ScoreExportError, match="无权导出"):
        await build_score_export_data(
            db, SimpleNamespace(id=2, role="teacher"), course_id=99
        )


@pytest.mark.asyncio
async def test_build_filters_by_class_and_course():
    class_a = SimpleNamespace(id=1, name="一班")
    math = SimpleNamespace(id=1, name="数学")
    exam_math = SimpleNamespace(id=1, course_id=1, title="期中考试", pass_score=60)
    student_a = SimpleNamespace(id=10, name="小明", class_id=1)
    db = FakeSession(
        [
            RowResult(
                [
                    (make_record(1, 10, 1, 80, "graded"), student_a, math, exam_math, class_a),
                ]
            ),
            ScalarResult([]),
        ]
    )

    data = await build_score_export_data(
        db, SimpleNamespace(id=3, role="admin"), class_id=1, course_id=1
    )

    assert len(data["student_scores"]) == 1
    assert data["student_scores"][0]["student_name"] == "小明"


@pytest.mark.asyncio
async def test_build_question_details_with_sort_order_numbering():
    class_a = SimpleNamespace(id=1, name="一班")
    math = SimpleNamespace(id=1, name="数学")
    exam = SimpleNamespace(id=1, course_id=1, title="期中考试", pass_score=60)
    student = SimpleNamespace(id=10, name="小明", class_id=1)
    record = make_record(1, 10, 1, 80, "graded")
    q_second = SimpleNamespace(
        id=2, exam_id=1, type="single", score=3, sort_order=2
    )
    q_first = SimpleNamespace(id=1, exam_id=1, type="essay", score=5, sort_order=1)
    db = FakeSession(
        [
            RowResult([(record, student, math, exam, class_a)]),
            RowResult(
                [
                    (SimpleNamespace(id=11, record_id=1, question_id=2, score=3), q_second),
                    (SimpleNamespace(id=12, record_id=1, question_id=1, score=4), q_first),
                ]
            ),
            ScalarResult([q_first, q_second]),
        ]
    )

    data = await build_score_export_data(db, SimpleNamespace(id=3, role="admin"))

    assert data["question_details"] == [
        {
            "student_name": "小明",
            "class_name": "一班",
            "course_name": "数学",
            "exam_title": "期中考试",
            "question_no": 1,
            "question_type": "简答",
            "score": 4,
            "full_score": 5,
        },
        {
            "student_name": "小明",
            "class_name": "一班",
            "course_name": "数学",
            "exam_title": "期中考试",
            "question_no": 2,
            "question_type": "单选",
            "score": 3,
            "full_score": 3,
        },
    ]


@pytest.mark.asyncio
async def test_build_question_details_numbers_by_full_question_list():
    class_a = SimpleNamespace(id=1, name="一班")
    math = SimpleNamespace(id=1, name="数学")
    exam = SimpleNamespace(id=1, course_id=1, title="期中考试", pass_score=60)
    student = SimpleNamespace(id=10, name="小明", class_id=1)
    record = make_record(1, 10, 1, 80, "graded")
    q_after_submission = SimpleNamespace(
        id=1, exam_id=1, type="judge", score=2, sort_order=1
    )
    q_answered = SimpleNamespace(
        id=2, exam_id=1, type="single", score=3, sort_order=2
    )
    db = FakeSession(
        [
            RowResult([(record, student, math, exam, class_a)]),
            RowResult(
                [
                    (SimpleNamespace(id=11, record_id=1, question_id=2, score=3), q_answered),
                ]
            ),
            ScalarResult([q_after_submission, q_answered]),
        ]
    )

    data = await build_score_export_data(db, SimpleNamespace(id=3, role="admin"))

    assert data["question_details"] == [
        {
            "student_name": "小明",
            "class_name": "一班",
            "course_name": "数学",
            "exam_title": "期中考试",
            "question_no": 2,
            "question_type": "单选",
            "score": 3,
            "full_score": 3,
        }
    ]


@pytest.mark.asyncio
async def test_admin_get_options_returns_all_classes_and_courses():
    db = FakeSession(
        [
            ScalarResult([SimpleNamespace(id=1, name="一班")]),
            ScalarResult([SimpleNamespace(id=1, name="数学")]),
        ]
    )

    options = await get_score_export_options(db, SimpleNamespace(id=3, role="admin"))

    assert options == {"classes": [{"id": 1, "name": "一班"}], "courses": [{"id": 1, "name": "数学"}]}


@pytest.mark.asyncio
async def test_teacher_get_options_limits_courses_to_assigned():
    db = FakeSession(
        [
            ScalarResult([SimpleNamespace(id=1, name="一班")]),
            ScalarResult([1]),
            ScalarResult(
                [
                    SimpleNamespace(id=1, name="数学"),
                ]
            ),
        ]
    )

    options = await get_score_export_options(db, SimpleNamespace(id=2, role="teacher"))

    assert options["classes"] == [{"id": 1, "name": "一班"}]
    assert options["courses"] == [{"id": 1, "name": "数学"}]


def test_render_creates_three_sheets_with_chinese_headers():
    datasets = {
        "class_summary": [
            {
                "class_name": "一班",
                "course_name": "数学",
                "exam_title": "期中考试",
                "student_count": 2,
                "avg_score": 67.5,
                "pass_rate": 50.0,
                "max_score": 80,
                "min_score": 55,
            }
        ],
        "student_scores": [
            {
                "student_name": "小明",
                "class_name": "一班",
                "course_name": "数学",
                "exam_title": "期中考试",
                "score": 80,
                "pass_score": 60,
                "status": "已阅卷",
            }
        ],
        "question_details": [
            {
                "student_name": "小明",
                "class_name": "一班",
                "course_name": "数学",
                "exam_title": "期中考试",
                "question_no": 1,
                "question_type": "简答",
                "score": 4,
                "full_score": 5,
            }
        ],
    }

    content, media_type, filename = render_score_export(datasets)

    assert media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert filename == "score-export.xlsx"
    workbook = load_workbook(BytesIO(content), read_only=True)
    assert workbook.sheetnames == ["班级成绩汇总", "学生成绩", "题目得分明细"]
    assert list(workbook["班级成绩汇总"].values) == [
        ("班级", "科目", "考试", "参考人数", "平均分", "及格率", "最高分", "最低分"),
        ("一班", "数学", "期中考试", 2, 67.5, 50.0, 80, 55),
    ]
    assert list(workbook["学生成绩"].values)[0] == (
        "学生姓名",
        "班级",
        "科目",
        "考试",
        "总分",
        "及格线",
        "状态",
    )
    assert list(workbook["题目得分明细"].values)[0] == (
        "学生姓名",
        "班级",
        "科目",
        "考试",
        "题号",
        "题型",
        "得分",
        "满分",
    )


def test_render_keeps_headers_when_datasets_empty():
    datasets = {
        "class_summary": [],
        "student_scores": [],
        "question_details": [],
    }

    content, _, _ = render_score_export(datasets)

    workbook = load_workbook(BytesIO(content), read_only=True)
    assert list(workbook["班级成绩汇总"].values) == [
        ("班级", "科目", "考试", "参考人数", "平均分", "及格率", "最高分", "最低分")
    ]
    assert workbook["学生成绩"].max_row == 1


@pytest.mark.asyncio
async def test_export_scores_endpoint_returns_streaming_response(monkeypatch):
    monkeypatch.setattr(
        statistics,
        "build_score_export_data",
        AsyncMock(
            return_value={
                "class_summary": [],
                "student_scores": [],
                "question_details": [],
            }
        ),
    )
    monkeypatch.setattr(
        statistics,
        "render_score_export",
        lambda datasets: (b"xlsx-content", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "score-export.xlsx"),
    )
    db = FakeSession([])

    response = await statistics.export_scores_file(
        class_id=None,
        course_id=None,
        db=db,
        current_user=SimpleNamespace(id=3, role="admin"),
    )

    assert isinstance(response, StreamingResponse)
    assert response.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert response.headers["content-disposition"].startswith("attachment;")
    assert '"score-export.xlsx"' in response.headers["content-disposition"]


@pytest.mark.asyncio
async def test_export_scores_endpoint_rejects_out_of_scope_course(monkeypatch):
    monkeypatch.setattr(
        statistics,
        "build_score_export_data",
        AsyncMock(side_effect=ScoreExportError("无权导出该科目的数据")),
    )
    db = FakeSession([])

    with pytest.raises(HTTPException) as exc_info:
        await statistics.export_scores_file(
            class_id=None,
            course_id=99,
            db=db,
            current_user=SimpleNamespace(id=2, role="teacher"),
        )

    assert exc_info.value.status_code == 400


@pytest.mark.asyncio
async def test_export_options_endpoint_returns_options(monkeypatch):
    monkeypatch.setattr(
        statistics,
        "get_score_export_options",
        AsyncMock(
            return_value={
                "classes": [{"id": 1, "name": "一班"}],
                "courses": [{"id": 1, "name": "数学"}],
            }
        ),
    )
    db = FakeSession([])

    response = await statistics.get_score_export_options_endpoint(
        db=db,
        current_user=SimpleNamespace(id=2, role="teacher"),
    )

    assert response["data"] == {
        "classes": [{"id": 1, "name": "一班"}],
        "courses": [{"id": 1, "name": "数学"}],
    }
